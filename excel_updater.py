import time
from openpyxl import load_workbook
from datetime import datetime
from phi.tools.yfinance import YFinanceTools
from financial_agent import analyze_stock_summary

EXCEL_FILE = "stocks.xlsx"
TICKER_COLUMN = "A"
ANALYSIS_COLUMN = "B"
TIMESTAMP_COLUMN = "C"

def get_news_timestamp(news_list):
    return news_list[0]['datetime'] if news_list else None

def update_excel():
    print("Checking Excel for updates...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):  # Skip header
        ticker_cell = row[0]
        analysis_cell = row[1]

        ticker = ticker_cell.value
        current_analysis = analysis_cell.value if analysis_cell else None

        should_update = True

        if should_update:
            print(f"Updating: {ticker}")
            analysis = analyze_stock_summary(ticker)
            analysis_cell.value = analysis

    wb.save(EXCEL_FILE)
    print("Excel updated.")

import xlwings as xw

def update_excel_live():
    wb = xw.Book(EXCEL_FILE)
    sheet = wb.sheets[0]

    yfinance = YFinanceTools(company_news=True)

    for row in range(2, sheet.cells.last_cell.row + 1):
        ticker = sheet.range(f"A{row}").value
        analysis = sheet.range(f"B{row}").value
        last_ts = sheet.range(f"C{row}").value
        if not ticker or str(ticker).strip() == "":
            break

        result = analyze_stock_summary(ticker)
        sheet.range(f"B{row}").value = result

    print("Live Excel update complete.")


# Run the loop
if __name__ == "__main__":
    while True:
        update_excel_live()
        time.sleep(600)  # every 10 minutes
